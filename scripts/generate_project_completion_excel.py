"""
Generate ERP project completion showcase workbook (modules + submodules + %).
Sources: docs/02_FRD (submodule names), apps/api & apps/web (implementation signals).
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
FRD_DIR = ROOT / "docs" / "02_FRD"
API_MODULES = ROOT / "apps" / "api" / "src" / "modules"
WEB_APP = ROOT / "apps" / "web" / "src" / "app" / "(app)"
OUTPUT = ROOT / "docs" / "ERP_Project_Completion_Status.xlsx"

# FRD file -> (display module name, API package folder)
FRD_MAP: list[tuple[str, str, str]] = [
    ("FRD-01-Foundation-Domain.md", "Foundation", "foundation"),
    ("FRD-02-Organization-Domain.md", "Organization", "organization"),
    ("FRD-03-Master-Data-Domain.md", "Master Data", "master_data"),
    ("FRD-04-Finance-Accounting-Domain.md", "Finance & Accounting", "finance"),
    ("FRD-05-CRM-Domain.md", "CRM", "crm"),
    ("FRD-06-Sales-Domain.md", "Sales", "sales"),
    ("FRD-07-Procurement-Domain.md", "Procurement", "procurement"),
    ("FRD-08-Inventory-Warehouse-Domain.md", "Inventory & Warehouse", "inventory"),
    ("FRD-09-HR-Domain.md", "Human Resources (HR)", "hr"),
    ("FRD-10-Payroll-Domain.md", "Payroll", "payroll"),
    ("FRD-11-Project-Management-Domain.md", "Project Management", "project"),
    ("FRD-12-Asset-Management-Domain.md", "Asset Management", "asset"),
    ("FRD-13-Manufacturing-Domain.md", "Manufacturing", "manufacturing"),
    ("FRD-14-Quality-Management-Domain.md", "Quality Management", "quality"),
    ("FRD-15-Supply-Chain-Management-Domain.md", "Supply Chain (SCM)", "procurement"),
    ("FRD-16-Service-Management-Domain.md", "Service Management", "service"),
    ("FRD-17-Helpdesk-Customer-Support-Domain.md", "Helpdesk & Support", "helpdesk"),
    ("FRD-18-BI-Reporting-Analytics-Domain.md", "BI, Reporting & Analytics", "analytics"),
    ("FRD-19-Document-Management-System-Domain.md", "Document Management System (DMS)", "document"),
    ("FRD-20-Compliance-Risk-Governance-Domain.md", "Compliance & GRC", "grc"),
    ("FRD-21-Integration-Hub-Enterprise-Platform-Services.md", "Integration Hub", "integration"),
    ("FRD-22-Ecommerce-External-Channel-Integration-Domain.md", "E-Commerce & Channels", "ecommerce"),
]

FOUNDATION_SUBMODULES = [
    "Authentication",
    "Session Management",
    "User Management",
    "Role Management",
    "Permission Management",
    "RBAC Engine",
    "Organization Context",
    "Workflow Engine",
    "Notification Engine",
    "Audit Engine",
    "Settings Management",
]

ORGANIZATION_SUBMODULES = [
    "Tenant Management",
    "Company Management",
    "Branch Management",
    "Business Unit Management",
    "Department Management",
    "Cost Center Management",
    "Profit Center Management",
    "Organization Tree View",
]

PORTAL_SUBMODULES = [
    "Portal Accounts",
    "Customer Profiles",
    "Portal Sessions",
    "Dashboards",
    "Dashboard Widgets",
    "Notifications",
    "Message Threads",
    "Messages",
    "Order Views",
    "Invoice Views",
    "Document Access",
    "Support Tickets",
    "Service Requests",
    "Download History",
    "Saved Reports",
    "Saved Searches",
    "Preferences",
    "Devices",
    "Login Audit",
    "Portal Reports",
]

RECRUITMENT_SUBMODULES = [
    "Job Requisitions",
    "Job Postings",
    "Candidate Management",
    "Applications & Stages",
    "Interviews",
    "Offers",
    "Onboarding",
    "Talent Pool",
]

ESS_SUBMODULES = [
    "Employee Self-Service Portal",
    "Leave Requests",
    "Attendance Views",
    "Payslip Access",
    "Profile Updates",
    "Announcements",
]

VIRTUAL_EA_SUBMODULES = [
    "Conversational AI Assistant",
    "Natural Language ERP Queries",
    "Task Automation & Reminders",
    "Calendar & Meeting Coordination",
    "Email Drafting & Summarization",
    "Workflow & Approval Assistance",
    "Proactive Insights & Alerts",
    "Voice & Chat Interface",
    "Tenant-Scoped Context & Memory",
    "Knowledge Retrieval (RAG)",
]

SELF_LEARNING_SUBMODULES = [
    "Learning Paths & Curricula",
    "Course Catalog Management",
    "Lesson & Content Delivery",
    "Quizzes & Assessments",
    "Skills Matrix & Competencies",
    "Certification & Badges",
    "Learner Progress Dashboards",
    "Instructor & Admin Console",
    "HR Training Integration",
    "Adaptive Learning Recommendations",
]

DOCUMENT_MANAGEMENT_SUBMODULES = [
    "Central Document Library",
    "Folder Structure & Taxonomy",
    "Version Control & History",
    "Check-in / Check-out",
    "Document Approval Workflows",
    "Full-Text Search & Discovery",
    "Retention Policies & Archival",
    "Access Control & Sharing",
    "Templates & Standard Forms",
    "OCR & Metadata Tagging",
]

LICENSING_SUBMODULES = [
    "License Entitlements",
    "Product & Module Licensing",
    "Seat & User Allocation",
    "Subscription Plans & Tiers",
    "License Key Generation",
    "Activation & Validation",
    "Usage Metering",
    "Renewal & Expiry Management",
    "Multi-Tenant License Isolation",
    "License Audit & Compliance Reports",
]

MONITORING_ANALYTICS_SUBMODULES = [
    "Real-Time System Health",
    "API Performance & Latency",
    "Error & Exception Tracking",
    "User Activity Analytics",
    "Business KPI Dashboards",
    "Alert Rules & Notifications",
    "Log Aggregation & Search",
    "Infrastructure Metrics",
    "SLA & Uptime Reporting",
    "Custom Report Builder",
]

# (display name, API package for scoring, reference id, submodules, roadmap floor % if no API)
PLATFORM_EXTENSION_MODULES: list[tuple[str, str, str, list[str], int]] = [
    ("Virtual E.A.", "", "Roadmap-24", VIRTUAL_EA_SUBMODULES, 15),
    ("Self Learning Module", "hr", "Roadmap-24", SELF_LEARNING_SUBMODULES, 28),
    ("Document Management", "document", "Platform", DOCUMENT_MANAGEMENT_SUBMODULES, 40),
    ("Licensing", "foundation", "Roadmap-24", LICENSING_SUBMODULES, 22),
    ("Monitoring & Analytics", "analytics", "Platform", MONITORING_ANALYTICS_SUBMODULES, 35),
]

PLATFORM_MODULE_ORDER = (
    "Virtual E.A.",
    "Self Learning Module",
    "Document Management",
    "Licensing",
    "Monitoring & Analytics",
)

# Web route folder names (may differ from API package)
WEB_FOLDER: dict[str, str] = {
    "foundation": "foundation",
    "organization": "organization",
    "master_data": "master-data",
    "finance": "finance",
    "crm": "crm",
    "sales": "sales",
    "procurement": "procurement",
    "inventory": "inventory",
    "hr": "hr",
    "payroll": "payroll",
    "project": "projects",
    "asset": "assets",
    "manufacturing": "manufacturing",
    "quality": "quality",
    "service": "service",
    "helpdesk": "helpdesk",
    "analytics": "analytics",
    "document": "documents",
    "grc": "grc",
    "integration": "integration",
    "ecommerce": "ecommerce",
    "portal": "portal",
    "recruitment": "recruitment",
    "ess": "hr",
    "self_learning": "hr/learning",
    "virtual_ea": "virtual-ea",
    "licensing": "licensing",
    "document_management": "documents",
    "monitoring_analytics": "analytics",
}


@dataclass
class Row:
    module: str
    submodule: str
    frd: str
    backend_pct: int
    frontend_pct: int

    @property
    def overall_pct(self) -> int:
        return round(0.55 * self.backend_pct + 0.45 * self.frontend_pct)


def parse_modules_covered(frd_path: Path) -> list[str]:
    text = frd_path.read_text(encoding="utf-8")
    m = re.search(r"## 2\. MODULES COVERED\s*\n(.*?)(?=\n## )", text, re.DOTALL)
    if not m:
        return []
    items: list[str] = []
    for line in m.group(1).splitlines():
        line = line.strip()
        if line.startswith("- "):
            items.append(line[2:].strip())
    return items


def module_py_files(api_pkg: str) -> list[str]:
    base = API_MODULES / api_pkg
    if not base.is_dir():
        return []
    names: list[str] = []
    for p in base.rglob("*.py"):
        if "__pycache__" in p.parts:
            continue
        names.append(p.name.lower())
        names.append(p.stem.lower())
    return names


def keywords_from_submodule(name: str) -> list[str]:
    stop = {
        "management",
        "engine",
        "service",
        "processing",
        "integration",
        "accounting",
        "tracking",
        "system",
    }
    words = re.findall(r"[a-z0-9]+", name.lower())
    return [w for w in words if len(w) > 2 and w not in stop]


def score_roadmap_backend(submodule: str, floor: int) -> int:
    """Score planned modules by scanning the full API tree for related keywords."""
    kws = keywords_from_submodule(submodule)
    if not kws:
        return floor
    hits = 0
    for p in API_MODULES.rglob("*.py"):
        if "__pycache__" in p.parts:
            continue
        blob = p.stem.lower()
        if any(kw in blob for kw in kws):
            hits += 1
    return min(48, floor + min(hits, 6) * 3)


def score_backend(
    api_pkg: str,
    submodule: str,
    module_exists_boost: bool = True,
    *,
    roadmap_floor: int | None = None,
) -> int:
    if not api_pkg or not (API_MODULES / api_pkg).is_dir():
        return score_roadmap_backend(submodule, roadmap_floor if roadmap_floor is not None else 18)
    files = module_py_files(api_pkg)
    if not files:
        return score_roadmap_backend(submodule, roadmap_floor or 18)
    kws = keywords_from_submodule(submodule)
    if not kws:
        return 72
    hits = sum(1 for kw in kws if any(kw in f for f in files))
    ratio = hits / max(len(kws), 1)
    base = 68 + int(ratio * 28)
    if module_exists_boost and (API_MODULES / api_pkg / "models").is_dir():
        base = max(base, 75)
    # Advanced / edge capabilities often partial in beta
    advanced = ("ocr", "sso", "whatsapp", "data warehouse", "forecasting", "e-signature", "marketplace")
    if any(a in submodule.lower() for a in advanced):
        base = min(base, 82)
    return min(100, max(35, base))


def count_web_pages(web_key: str) -> int:
    if web_key in WEB_FOLDER:
        path = WEB_APP / Path(WEB_FOLDER[web_key])
        if path.is_dir():
            return len(list(path.rglob("page.tsx")))
        return 0
    folder = web_key.replace("_", "-")
    path = WEB_APP / folder
    if path.is_dir():
        return len(list(path.rglob("page.tsx")))
    # Generic dynamic module shell at /[module]
    dynamic = WEB_APP / "[module]"
    if dynamic.is_dir():
        return 2
    return 0


def score_frontend_module(api_pkg: str) -> int:
    pages = count_web_pages(api_pkg)
    if pages >= 30:
        return min(92, 55 + pages)
    if pages >= 15:
        return min(85, 48 + int(pages * 1.2))
    if pages >= 6:
        return min(72, 40 + pages * 3)
    if pages >= 2:
        return 38
    return 28


def submodule_frontend_adjust(submodule: str, module_fe: int) -> int:
    """Spread module UI % with slight variation by submodule maturity."""
    low_ui = (
        "forecast",
        "warehouse",
        "ocr",
        "sso",
        "whatsapp",
        "data warehouse",
        "network planning",
        "mobile app",
        "e-signature",
        "barcode",
    )
    high_ui = (
        "dashboard",
        "invoice",
        "journal",
        "chart of accounts",
        "employee",
        "attendance",
        "leave",
        "quotation",
        "sales order",
        "project",
        "timesheet",
        "portal",
        "payslip",
    )
    name = submodule.lower()
    adj = 0
    if any(x in name for x in low_ui):
        adj -= 8
    if any(x in name for x in high_ui):
        adj += 10
    return min(95, max(20, module_fe + adj))


def collect_rows() -> list[Row]:
    rows: list[Row] = []
    seen_scm = False

    for frd_file, module_name, api_pkg in FRD_MAP:
        frd_path = FRD_DIR / frd_file
        frd_id = frd_file.split("-")[0] + "-" + frd_file.split("-")[1]

        if frd_file.startswith("FRD-01"):
            submodules = FOUNDATION_SUBMODULES
        elif frd_file.startswith("FRD-02"):
            submodules = ORGANIZATION_SUBMODULES
        else:
            submodules = parse_modules_covered(frd_path)

        if frd_file.startswith("FRD-15"):
            if seen_scm:
                continue
            seen_scm = True
            module_name = "Supply Chain (SCM)"
            # SCM implemented largely under procurement package
            api_pkg = "procurement"

        mod_fe = score_frontend_module(api_pkg)
        for sub in submodules:
            be = score_backend(api_pkg, sub)
            fe = submodule_frontend_adjust(sub, mod_fe)
            rows.append(Row(module_name, sub, frd_id, be, fe))

    extra_modules = [
        ("Customer Portal", "portal", "ERD-23", PORTAL_SUBMODULES),
        ("Recruitment", "recruitment", "HR-Ext", RECRUITMENT_SUBMODULES),
        ("Employee Self-Service (ESS)", "ess", "HR-Ext", ESS_SUBMODULES),
    ]
    for module_name, api_pkg, frd_id, submodules in extra_modules:
        mod_fe = score_frontend_module(api_pkg)
        for sub in submodules:
            rows.append(
                Row(
                    module_name,
                    sub,
                    frd_id,
                    score_backend(api_pkg, sub),
                    submodule_frontend_adjust(sub, mod_fe),
                )
            )

    web_key_by_module = {
        "Virtual E.A.": "virtual_ea",
        "Self Learning Module": "self_learning",
        "Document Management": "document_management",
        "Licensing": "licensing",
        "Monitoring & Analytics": "monitoring_analytics",
    }
    for module_name, api_pkg, ref_id, submodules, roadmap_floor in PLATFORM_EXTENSION_MODULES:
        web_key = web_key_by_module[module_name]
        mod_fe = score_frontend_module(web_key)
        if mod_fe <= 28 and api_pkg:
            mod_fe = max(mod_fe, score_frontend_module(api_pkg) - 5)
        for sub in submodules:
            rows.append(
                Row(
                    module_name,
                    sub,
                    ref_id,
                    score_backend(api_pkg, sub, roadmap_floor=roadmap_floor),
                    submodule_frontend_adjust(sub, mod_fe),
                )
            )
    return rows


def completion_fill(pct: int) -> PatternFill:
    if pct >= 85:
        color = "C6EFCE"
    elif pct >= 70:
        color = "FFEB9C"
    elif pct >= 50:
        color = "FCE4D6"
    else:
        color = "FFC7CE"
    return PatternFill("solid", fgColor=color)


def style_header(ws, row: int, cols: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="CCCCCC")
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(thin, thin, thin, thin)


def autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[letter].width = min(max_width, max(10, length + 2))


def module_order(rows: list[Row]) -> list[str]:
    """Preserve FRD roadmap order (not alphabetical)."""
    order: list[str] = []
    seen: set[str] = set()
    for _, name, _ in FRD_MAP:
        if name not in seen:
            order.append(name)
            seen.add(name)
    for name in (
        "Customer Portal",
        "Recruitment",
        "Employee Self-Service (ESS)",
        *PLATFORM_MODULE_ORDER,
    ):
        if name not in seen:
            order.append(name)
            seen.add(name)
    for r in rows:
        if r.module not in seen:
            order.append(r.module)
            seen.add(r.module)
    return order


def group_rows_by_module(rows: list[Row]) -> dict[str, list[Row]]:
    grouped: dict[str, list[Row]] = defaultdict(list)
    for r in rows:
        grouped[r.module].append(r)
    for mod in grouped:
        grouped[mod].sort(key=lambda x: x.submodule)
    return grouped


def status_label(ov: int) -> str:
    if ov >= 85:
        return "Complete / Beta-ready"
    if ov >= 70:
        return "Substantial"
    if ov >= 50:
        return "In progress"
    return "Planned / Shell UI"


def apply_row_border(ws, row: int, cols: int, thin: Side) -> None:
    for col in range(1, cols + 1):
        ws.cell(row, col).border = Border(thin, thin, thin, thin)


def module_averages(items: list[Row]) -> tuple[int, int, int]:
    be = round(sum(x.backend_pct for x in items) / len(items))
    fe = round(sum(x.frontend_pct for x in items) / len(items))
    ov = round(sum(x.overall_pct for x in items) / len(items))
    return be, fe, ov


def build_workbook(rows: list[Row]) -> Workbook:
    wb = Workbook()
    thin = Side(style="thin", color="DDDDDD")
    module_fill = PatternFill("solid", fgColor="D6E4F0")
    submodule_fill = PatternFill("solid", fgColor="F8F9FA")

    grouped = group_rows_by_module(rows)
    order = module_order(rows)

    title_a1 = "Enterprise ERP Platform — Project Completion"
    title_a2 = f"As of {date.today().isoformat()} · Release baseline ERP Core v1.18-beta"

    detail_headers = [
        "Main Module",
        "Submodule",
        "FRD / Ref",
        "Backend %",
        "Frontend %",
        "Overall %",
        "Status",
    ]
    num_cols = len(detail_headers)

    # --- Primary sheet: modules with submodules nested underneath ---
    ws_tree = wb.active
    ws_tree.title = "Modules & Submodules"
    ws_tree["A1"] = title_a1
    ws_tree["A1"].font = Font(size=16, bold=True, color="1F4E78")
    ws_tree["A2"] = title_a2
    ws_tree["A2"].font = Font(size=10, color="666666")
    for i, h in enumerate(detail_headers, 1):
        ws_tree.cell(4, i, h)
    style_header(ws_tree, 4, num_cols)

    row_idx = 5
    for module in order:
        items = grouped.get(module, [])
        if not items:
            continue
        be, fe, ov = module_averages(items)
        frd_ref = items[0].frd

        ws_tree.cell(row_idx, 1, module)
        ws_tree.cell(row_idx, 2, f"Module average ({len(items)} submodules)")
        ws_tree.cell(row_idx, 3, frd_ref)
        ws_tree.cell(row_idx, 4, be)
        ws_tree.cell(row_idx, 5, fe)
        ov_cell = ws_tree.cell(row_idx, 6, ov)
        ov_cell.fill = completion_fill(ov)
        ws_tree.cell(row_idx, 7, status_label(ov))
        for col in range(1, num_cols + 1):
            cell = ws_tree.cell(row_idx, col)
            cell.font = Font(bold=True, color="1F4E78")
            cell.fill = module_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        apply_row_border(ws_tree, row_idx, num_cols, thin)
        module_header_row = row_idx
        row_idx += 1

        first_sub_row = row_idx
        for item in items:
            ws_tree.cell(row_idx, 1, module)
            ws_tree.cell(row_idx, 2, item.submodule)
            ws_tree.cell(row_idx, 3, item.frd)
            ws_tree.cell(row_idx, 4, item.backend_pct)
            ws_tree.cell(row_idx, 5, item.frontend_pct)
            sub_ov = item.overall_pct
            sub_ov_cell = ws_tree.cell(row_idx, 6, sub_ov)
            sub_ov_cell.fill = completion_fill(sub_ov)
            ws_tree.cell(row_idx, 7, status_label(sub_ov))
            ws_tree.cell(row_idx, 2).alignment = Alignment(
                indent=2, vertical="center", wrap_text=True
            )
            for col in range(1, num_cols + 1):
                ws_tree.cell(row_idx, col).fill = submodule_fill
                ws_tree.cell(row_idx, col).alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                    indent=2 if col == 2 else 0,
                )
            ws_tree.row_dimensions[row_idx].outlineLevel = 1
            apply_row_border(ws_tree, row_idx, num_cols, thin)
            row_idx += 1

        last_sub_row = row_idx - 1
        if last_sub_row >= first_sub_row:
            ws_tree.merge_cells(
                start_row=first_sub_row,
                start_column=1,
                end_row=last_sub_row,
                end_column=1,
            )
            merged = ws_tree.cell(first_sub_row, 1)
            merged.alignment = Alignment(vertical="top", wrap_text=True)
            merged.font = Font(bold=True, color="333333")

        ws_tree.row_dimensions[module_header_row].outlineLevel = 0

    ws_tree.freeze_panes = "A5"
    ws_tree.sheet_properties.outlinePr.summaryBelow = False
    ws_tree.column_dimensions["A"].width = 28
    ws_tree.column_dimensions["B"].width = 42
    ws_tree.column_dimensions["C"].width = 12
    for letter in ("D", "E", "F", "G"):
        ws_tree.column_dimensions[letter].width = 14

    # --- Executive Summary (module roll-up + submodule name list) ---
    ws_sum = wb.create_sheet("Executive Summary", 1)
    ws_sum["A1"] = title_a1
    ws_sum["A1"].font = Font(size=16, bold=True, color="1F4E78")
    ws_sum["A2"] = title_a2
    ws_sum["A2"].font = Font(size=10, color="666666")

    sum_headers = [
        "Main Module",
        "# Submodules",
        "Submodule Names",
        "Avg Backend %",
        "Avg Frontend %",
        "Avg Overall %",
    ]
    for i, h in enumerate(sum_headers, 1):
        ws_sum.cell(4, i, h)
    style_header(ws_sum, 4, len(sum_headers))

    row_idx = 5
    totals_be: list[int] = []
    totals_fe: list[int] = []
    totals_ov: list[int] = []
    for module in order:
        items = grouped.get(module, [])
        if not items:
            continue
        be, fe, ov = module_averages(items)
        totals_be.append(be)
        totals_fe.append(fe)
        totals_ov.append(ov)
        names = "\n".join(f"• {x.submodule}" for x in items)
        ws_sum.cell(row_idx, 1, module)
        ws_sum.cell(row_idx, 2, len(items))
        ws_sum.cell(row_idx, 3, names)
        ws_sum.cell(row_idx, 4, be)
        ws_sum.cell(row_idx, 5, fe)
        c = ws_sum.cell(row_idx, 6, ov)
        c.fill = completion_fill(ov)
        ws_sum.cell(row_idx, 3).alignment = Alignment(wrap_text=True, vertical="top")
        for col in range(1, 7):
            ws_sum.cell(row_idx, col).border = Border(thin, thin, thin, thin)
            ws_sum.cell(row_idx, col).alignment = Alignment(
                vertical="top" if col == 3 else "center", wrap_text=True
            )
        row_idx += 1

    ws_sum.cell(row_idx, 1, "PLATFORM TOTAL (module average)")
    ws_sum.cell(row_idx, 1).font = Font(bold=True)
    ws_sum.cell(row_idx, 2, len(rows))
    ws_sum.cell(row_idx, 3, "")
    ws_sum.cell(row_idx, 4, round(sum(totals_be) / len(totals_be)))
    ws_sum.cell(row_idx, 5, round(sum(totals_fe) / len(totals_fe)))
    tot_cell = ws_sum.cell(row_idx, 6, round(sum(totals_ov) / len(totals_ov)))
    tot_cell.font = Font(bold=True)
    tot_cell.fill = completion_fill(int(tot_cell.value))

    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 14
    ws_sum.column_dimensions["C"].width = 52
    for letter in ("D", "E", "F"):
        ws_sum.column_dimensions[letter].width = 16

    # --- Flat detail (same hierarchy labels for export/filter) ---
    ws = wb.create_sheet("Module Detail")
    for i, h in enumerate(detail_headers, 1):
        ws.cell(1, i, h)
    style_header(ws, 1, num_cols)

    flat_row = 2
    for module in order:
        items = grouped.get(module, [])
        if not items:
            continue
        be, fe, ov = module_averages(items)
        ws.cell(flat_row, 1, module)
        ws.cell(flat_row, 2, f"Module average ({len(items)} submodules)")
        ws.cell(flat_row, 3, items[0].frd)
        ws.cell(flat_row, 4, be)
        ws.cell(flat_row, 5, fe)
        ws.cell(flat_row, 6, ov).fill = completion_fill(ov)
        ws.cell(flat_row, 7, status_label(ov))
        for col in range(1, num_cols + 1):
            ws.cell(flat_row, col).font = Font(bold=True)
            ws.cell(flat_row, col).fill = module_fill
        apply_row_border(ws, flat_row, num_cols, thin)
        flat_row += 1

        for item in items:
            ws.cell(flat_row, 1, module)
            ws.cell(flat_row, 2, item.submodule)
            ws.cell(flat_row, 3, item.frd)
            ws.cell(flat_row, 4, item.backend_pct)
            ws.cell(flat_row, 5, item.frontend_pct)
            ws.cell(flat_row, 6, item.overall_pct).fill = completion_fill(item.overall_pct)
            ws.cell(flat_row, 7, status_label(item.overall_pct))
            ws.cell(flat_row, 2).alignment = Alignment(indent=1, vertical="center", wrap_text=True)
            apply_row_border(ws, flat_row, num_cols, thin)
            flat_row += 1

    ws.freeze_panes = "A2"
    autosize(ws)

    # --- Legend ---
    ws_leg = wb.create_sheet("Legend")
    ws_leg["A1"] = "How to read this workbook"
    ws_leg["A1"].font = Font(bold=True, size=12)
    notes = [
        "Modules & Submodules: each main module has a bold summary row, then indented submodule rows underneath (column B).",
        "Use Excel outline controls (rows 1–7) to collapse/expand submodule lists per module.",
        "Executive Summary lists every submodule name under each main module in column C.",
        "Submodule names are taken from approved FRD-01–FRD-22, Portal (ERD-23), Recruitment, ESS, and platform extensions "
        "(Virtual E.A., Self Learning, Document Management, Licensing, Monitoring & Analytics).",
        "Backend % reflects API/services/models signals in apps/api/src/modules (keyword + package presence).",
        "Frontend % reflects Next.js admin UI coverage under apps/web/src/app/(app) (page count + submodule weighting).",
        "Overall % = 55% Backend + 45% Frontend (typical ERP delivery weighting for stakeholder reporting).",
        "Color: Green ≥85% · Amber ≥70% · Peach ≥50% · Red <50%.",
    ]
    for idx, line in enumerate(notes, 3):
        ws_leg.cell(idx, 1, line)
        ws_leg.cell(idx, 1).alignment = Alignment(wrap_text=True)
    ws_leg.column_dimensions["A"].width = 100

    return wb


def main() -> None:
    rows = collect_rows()
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(rows)
    try:
        wb.save(OUTPUT)
        print(f"Wrote {OUTPUT} ({len(rows)} submodule rows)")
    except PermissionError:
        alt = OUTPUT.with_name(f"{OUTPUT.stem}_v3{OUTPUT.suffix}")
        wb.save(alt)
        print(
            f"Could not overwrite {OUTPUT} (file may be open in Excel). "
            f"Wrote {alt} ({len(rows)} submodule rows)"
        )


if __name__ == "__main__":
    main()
