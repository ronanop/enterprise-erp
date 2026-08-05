"""Heuristic extraction of vendor invoice fields from PDF/text/Excel."""

from __future__ import annotations

import base64
import binascii
import re
from datetime import date, datetime
from io import BytesIO
from typing import Any


def _decode_pdf_parentheses(raw: bytes) -> str:
    parts: list[str] = []
    for match in re.finditer(rb"\((?:\\.|[^\\)])*\)", raw):
        chunk = match.group(0)[1:-1]
        chunk = chunk.replace(rb"\(", b"(").replace(rb"\)", b")").replace(rb"\\", b"\\")
        parts.append(chunk.decode("latin-1", errors="ignore"))
    if parts:
        return " ".join(parts)
    runs = re.findall(rb"[\x20-\x7e]{4,}", raw)
    return " ".join(r.decode("ascii", errors="ignore") for r in runs[:4000])


def _text_from_pdf_pypdf(raw: bytes) -> str | None:
    try:
        from pypdf import PdfReader

        reader = PdfReader(BytesIO(raw))
        chunks: list[str] = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                chunks.append(text)
        joined = "\n".join(chunks).strip()
        return joined if joined else None
    except Exception:
        return None


def _text_from_excel(raw: bytes) -> str | None:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        parts: list[str] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if cells:
                    parts.append(" ".join(cells))
        joined = "\n".join(parts).strip()
        return joined if joined else None
    except Exception:
        return None


def _text_from_xls(raw: bytes) -> str | None:
    try:
        import xlrd

        book = xlrd.open_workbook(file_contents=raw)
        parts: list[str] = []
        for sheet in book.sheets():
            for row_idx in range(sheet.nrows):
                cells = [
                    str(sheet.cell_value(row_idx, col_idx)).strip()
                    for col_idx in range(sheet.ncols)
                    if str(sheet.cell_value(row_idx, col_idx)).strip()
                ]
                if cells:
                    parts.append(" ".join(cells))
        joined = "\n".join(parts).strip()
        return joined if joined else None
    except Exception:
        return None


def _text_from_bytes(raw: bytes, file_name: str) -> str:
    lower = file_name.lower()
    if lower.endswith((".txt", ".csv")):
        return raw.decode("utf-8", errors="ignore")
    if lower.endswith((".xlsx", ".xlsm")):
        excel_text = _text_from_excel(raw)
        if excel_text:
            return excel_text
    if lower.endswith(".xls"):
        xls_text = _text_from_xls(raw)
        if xls_text:
            return xls_text
    if lower.endswith(".pdf") or raw[:4] == b"%PDF":
        pypdf_text = _text_from_pdf_pypdf(raw)
        if pypdf_text:
            return pypdf_text
        if raw[:4] != b"%PDF":
            plain = raw.decode("utf-8", errors="ignore").strip()
            if plain:
                return plain
        return _decode_pdf_parentheses(raw)
    if lower.endswith(
        (".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".tif", ".tiff", ".heic", ".heif")
    ):
        return ""
    if lower.endswith((".doc", ".docx")):
        return raw.decode("utf-8", errors="ignore")
    return raw.decode("utf-8", errors="ignore")


def _parse_amount(token: str) -> float | None:
    cleaned = token.strip().replace(",", "")
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _parse_date(token: str) -> date | None:
    token = token.strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d %b %Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(token, fmt).date()
        except ValueError:
            continue
    return None


def _find_date(text: str) -> date | None:
    patterns = [
        r"invoice\s*date\s*[:\-]?\s*(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})",
        r"bill\s*date\s*[:\-]?\s*(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})",
        r"date\s*of\s*invoice\s*[:\-]?\s*(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})",
        r"date\s*[:\-]?\s*(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})",
        r"(\d{1,2}[-/]\d{1,2}[-/]\d{4})",
    ]
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            parsed = _parse_date(m.group(1))
            if parsed:
                return parsed
    return None


def _find_invoice_number(text: str) -> str | None:
    patterns = [
        r"invoice\s*(?:no\.?|number|#)\s*[:\-]?\s*([A-Za-z0-9][A-Za-z0-9/_\-.]{2,40})",
        r"inv\s*(?:no\.?|#)\s*[:\-]?\s*([A-Za-z0-9][A-Za-z0-9/_\-.]{2,40})",
        r"bill\s*(?:no\.?|#)\s*[:\-]?\s*([A-Za-z0-9][A-Za-z0-9/_\-.]{2,40})",
        r"tax\s*invoice\s*(?:no\.?|#)?\s*[:\-]?\s*([A-Za-z0-9][A-Za-z0-9/_\-.]{2,40})",
    ]
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return None


def _looks_like_money(value: float, raw_token: str = "") -> bool:
    if raw_token and "," in raw_token:
        return True
    if value >= 100:
        return True
    if abs(value - round(value)) > 0.009:
        return True
    return False


def _numbers_on_line(line: str) -> list[tuple[float, str]]:
    found: list[tuple[float, str]] = []
    for m in re.finditer(r"(?<!\d)([\d,]+\.?\d*)(?!\d)", line):
        val = _parse_amount(m.group(1))
        if val is not None:
            found.append((val, m.group(1)))
    return found


def _find_qty_amount_pair(text: str) -> tuple[float | None, float | None]:
    """When labels share a line (common in PDF tables), pick qty vs amount by position/heuristics."""
    for line in text.splitlines():
        lower = line.lower()
        if "quantity" not in lower and "qty" not in lower:
            continue
        if "total amount" not in lower and "amount" not in lower:
            continue
        nums = _numbers_on_line(line)
        if not nums:
            continue
        if len(nums) >= 2:
            last_val, last_raw = nums[-1]
            if _looks_like_money(last_val, last_raw):
                for i in range(len(nums) - 2, -1, -1):
                    qty_val, _ = nums[i]
                    if _looks_like_quantity(qty_val) and qty_val < last_val:
                        return qty_val, last_val
            qty_val, qty_raw = nums[0]
            amount_candidates = [
                (v, r) for v, r in nums[1:] if _looks_like_money(v, r) or v >= qty_val
            ]
            if amount_candidates:
                best = max(amount_candidates, key=lambda t: t[0])
                if _looks_like_quantity(qty_val) and best[0] > qty_val:
                    return qty_val, best[0]
        if len(nums) == 1:
            only_val, only_raw = nums[0]
            if _looks_like_money(only_val, only_raw):
                return None, only_val
            if _looks_like_quantity(only_val):
                return only_val, None
    return None, None


def _looks_like_quantity(value: float) -> bool:
    if value < 0:
        return False
    if value > 5000:
        return False
    # Large whole numbers are usually rate/amount, not PCS qty.
    if value >= 100 and abs(value - round(value)) < 0.0001:
        return False
    if abs(value - round(value)) < 0.0001:
        return True
    return value < 100


def _find_quantity(text: str, compact: str) -> float | None:
    qty_labels = [
        r"(?:total\s+)?(?:qty|quantity)\s*(?:nos?\.?|units?)?\s*[:\-]\s*([\d,]+\.?\d*)",
        r"(?:qty|quantity)\s*[:\-]?\s*([\d,]+\.?\d*)",
    ]
    for line in text.splitlines():
        line_stripped = line.strip()
        if not line_stripped:
            continue
        for pat in qty_labels:
            m = re.search(pat, line_stripped, re.IGNORECASE)
            if m:
                val = _parse_amount(m.group(1))
                if val is not None and _looks_like_quantity(val):
                    return val
    for pat in qty_labels:
        for m in re.finditer(pat, compact, re.IGNORECASE):
            val = _parse_amount(m.group(1))
            if val is not None and _looks_like_quantity(val):
                return val
    # Sum per-line qty in item tables (integer qty before rate/amount)
    line_qty: list[float] = []
    for m in re.finditer(
        r"(?:^|\s)(?:qty|quantity)\s*[:\-]?\s*([\d,]+\.?\d*)",
        compact,
        re.IGNORECASE,
    ):
        val = _parse_amount(m.group(1))
        if val is not None and _looks_like_quantity(val):
            line_qty.append(val)
    if line_qty:
        return sum(line_qty) if len(line_qty) > 1 else line_qty[0]
    return None


def _find_quantity_from_qty_pcs_table(text: str) -> float | None:
    """Line-item qty from tables headed Qty PCS / Qty + Rate (common on Indian tax invoices)."""
    lines = text.splitlines()
    header_idx: int | None = None
    for i, line in enumerate(lines):
        lower = line.lower()
        has_qty = re.search(r"\bqty\b", lower)
        if not has_qty:
            continue
        if (
            re.search(r"qty\s*\.?\s*pcs", lower)
            or re.search(r"\bpcs\b", lower)
            or re.search(r"\brate\b", lower)
            or "₹" in line
            or ("sac" in lower and "code" in lower)
        ):
            header_idx = i
            break
    if header_idx is None:
        return None

    qtys: list[float] = []
    for line in lines[header_idx + 1 : header_idx + 120]:
        stripped = line.strip()
        if not stripped:
            continue
        lower = stripped.lower()
        if any(
            k in lower
            for k in (
                "sub total",
                "subtotal",
                "grand total",
                "taxable value",
                "amount payable",
                "total amount",
                "igst",
                "cgst",
                "sgst",
            )
        ):
            break
        if re.fullmatch(r"\d+(?:\.\d+)?", stripped):
            val = float(stripped)
            if _looks_like_quantity(val):
                qtys.append(val)
            continue
        m = re.search(
            r"(?<![\d,.])([1-9]\d{0,4}(?:\.\d+)?)\s+(?:₹|rs\.?\s*)?[\d,]+(?:\.\d+)?",
            stripped,
            re.IGNORECASE,
        )
        if m:
            val = _parse_amount(m.group(1))
            if val is not None and _looks_like_quantity(val):
                qtys.append(val)
    if qtys:
        return sum(qtys)
    return None


def _find_quantity_after_hsn(compact: str) -> float | None:
    qtys: list[float] = []
    for m in re.finditer(
        r"(?:hsn|sac)(?:\s*/\s*sac)?\s*(?:code)?\s*[:\s]*\d{4,8}"
        r"[^\d]{0,240}?"
        r"([1-9]\d{0,4}(?:\.\d+)?)\s+(?:₹|rs\.?\s*)?[\d,]+(?:\.\d+)?",
        compact,
        re.IGNORECASE,
    ):
        val = _parse_amount(m.group(1))
        if val is not None and _looks_like_quantity(val):
            qtys.append(val)
    if qtys:
        return sum(qtys)
    return None


def _find_quantity_qty_before_rate(compact: str) -> float | None:
    qtys: list[float] = []
    for m in re.finditer(
        r"(?<![\d,.])([1-9]\d{0,4}(?:\.\d+)?)\s+(?:₹|rs\.?\s*)?[\d,]+(?:\.\d+)?",
        compact,
        re.IGNORECASE,
    ):
        val = _parse_amount(m.group(1))
        if val is not None and _looks_like_quantity(val):
            qtys.append(val)
    if qtys:
        return sum(qtys)
    return None


def _find_quantity_from_qty_pcs_compact(compact: str) -> float | None:
    m = re.search(r"qty\s*pcs", compact, re.IGNORECASE)
    if not m:
        return None
    tail = compact[m.end() : m.end() + 800]
    qtys: list[float] = []
    for hit in re.finditer(
        r"(?<![\d,.])([1-9]\d{0,4}(?:\.\d+)?)\s+(?:₹|rs\.?\s*)?[\d,]+(?:\.\d+)?",
        tail,
        re.IGNORECASE,
    ):
        val = _parse_amount(hit.group(1))
        if val is not None and _looks_like_quantity(val):
            qtys.append(val)
    if qtys:
        return sum(qtys)
    return None


def _resolve_invoice_quantity(
    text: str, compact: str, grid_qty: Any, pair_qty: float | None
) -> float | None:
    if grid_qty is not None:
        try:
            return float(grid_qty)
        except (TypeError, ValueError):
            pass
    for finder in (
        lambda: _find_quantity_from_qty_pcs_table(text),
        lambda: _find_quantity_from_qty_pcs_compact(compact),
        lambda: _find_quantity_after_hsn(compact),
        lambda: _find_quantity_qty_before_rate(compact),
        lambda: _find_quantity(text, compact),
    ):
        val = finder()
        if val is not None:
            return float(val)
    if pair_qty is not None and _looks_like_quantity(float(pair_qty)):
        return float(pair_qty)
    return None


def _find_subtotal(compact: str, exclude: set[float]) -> float | None:
    patterns = [
        r"(?:sub\s*total|subtotal|taxable\s*(?:value|amount)|amount\s*before\s*tax|total\s*(?:excl\.?|excluding)\s*tax)[^\d]{0,40}([\d,]+\.?\d*)",
        r"(?:total\s*amount\s*without\s*tax|value\s*without\s*tax)[^\d]{0,40}([\d,]+\.?\d*)",
        r"(?:net\s*amount|taxable\s*value)[^\d]{0,40}([\d,]+\.?\d*)",
    ]
    candidates: list[float] = []
    for pat in patterns:
        for m in re.finditer(pat, compact, re.IGNORECASE):
            raw = m.group(1)
            val = _parse_amount(raw)
            if val is None or any(abs(val - ex) < 0.01 for ex in exclude):
                continue
            if _looks_like_money(val, raw) or val >= 50:
                candidates.append(val)
    if candidates:
        return max(candidates)
    return None


def _find_total_amount(compact: str, exclude: set[float]) -> float | None:
    patterns = [
        r"(?:grand\s+total|total\s+invoice\s+value|amount\s+payable|invoice\s+total)[^\d]{0,40}([\d,]+\.?\d*)",
        r"total\s+amount(?!\s*(?:without|excl))[^\d]{0,40}([\d,]+\.?\d*)",
        r"invoice\s+amount[^\d]{0,40}([\d,]+\.?\d*)",
    ]
    candidates: list[float] = []
    for pat in patterns:
        for m in re.finditer(pat, compact, re.IGNORECASE):
            raw = m.group(1)
            val = _parse_amount(raw)
            if val is None or any(abs(val - ex) < 0.01 for ex in exclude):
                continue
            if _looks_like_money(val, raw) or val >= 50:
                candidates.append(val)
    if not candidates:
        return None
    return max(candidates)


def _find_largest_money_fallback(compact: str, exclude: set[float]) -> float | None:
    candidates: list[float] = []
    for m in re.finditer(r"(?<!\d)([\d]{1,3}(?:,\d{2,3})+(?:\.\d{1,2})?|\d+\.\d{2})(?!\d)", compact):
        val = _parse_amount(m.group(1))
        if val is None or val < 10:
            continue
        if any(abs(val - ex) < 0.01 for ex in exclude):
            continue
        candidates.append(val)
    if not candidates:
        for m in re.finditer(r"(?<!\d)([\d]{4,}(?:,\d{3})*)(?!\d)", compact):
            val = _parse_amount(m.group(1))
            if val is None or val < 100:
                continue
            if any(abs(val - ex) < 0.01 for ex in exclude):
                continue
            candidates.append(val)
    return max(candidates) if candidates else None


def _extract_from_excel_grid(raw: bytes) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}

    workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    out: dict[str, Any] = {}
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header_idx = None
        col_map: dict[str, int] = {}
        for idx, row in enumerate(rows[:15]):
            cells = [str(c).strip().lower() if c is not None else "" for c in row]
            if not any(cells):
                continue
            hits = 0
            for i, cell in enumerate(cells):
                if ("invoice" in cell and "no" in cell) or cell in ("inv no", "invoice number"):
                    col_map["invoice_no"] = i
                    hits += 1
                if "invoice" in cell and "date" in cell:
                    col_map["invoice_date"] = i
                    hits += 1
                if cell in ("qty", "quantity", "total qty") or cell.startswith("qty") or "qty" in cell and "pcs" in cell:
                    col_map["quantity"] = i
                    hits += 1
                if "total" in cell and "amount" in cell:
                    col_map["amount"] = i
                    hits += 1
                elif cell in ("subtotal", "sub total", "taxable value", "net amount"):
                    col_map["amount"] = i
                    hits += 1
                elif cell == "amount" and "amount" not in col_map:
                    col_map["amount"] = i
                    hits += 1
            if hits >= 2:
                header_idx = idx
                break
        if header_idx is None:
            continue
        qty_sum = 0.0
        qty_count = 0
        amount_vals: list[float] = []
        for row in rows[header_idx + 1 : header_idx + 200]:
            if not row:
                continue
            if "invoice_no" in col_map and not out.get("vendor_invoice_number"):
                raw_no = row[col_map["invoice_no"]]
                if raw_no is not None and str(raw_no).strip():
                    out["vendor_invoice_number"] = str(raw_no).strip()
            if "invoice_date" in col_map and not out.get("vendor_invoice_date"):
                raw_d = row[col_map["invoice_date"]]
                if raw_d is not None:
                    if isinstance(raw_d, datetime):
                        out["vendor_invoice_date"] = raw_d.date().isoformat()
                    elif hasattr(raw_d, "isoformat"):
                        out["vendor_invoice_date"] = raw_d.isoformat()[:10]
                    else:
                        parsed = _parse_date(str(raw_d))
                        if parsed:
                            out["vendor_invoice_date"] = parsed.isoformat()
            if "quantity" in col_map:
                raw_q = row[col_map["quantity"]]
                if raw_q is not None:
                    try:
                        q = float(raw_q)
                        if _looks_like_quantity(q):
                            qty_sum += q
                            qty_count += 1
                    except (TypeError, ValueError):
                        pass
            if "amount" in col_map:
                raw_a = row[col_map["amount"]]
                if raw_a is not None:
                    try:
                        amount_vals.append(float(raw_a))
                    except (TypeError, ValueError):
                        pass
        if qty_count > 0:
            out["vendor_invoice_quantity"] = qty_sum
        if amount_vals:
            money_vals = [
                v
                for v in amount_vals
                if _looks_like_money(v) or v >= max(qty_sum, 50)
            ]
            pick_from = money_vals if money_vals else amount_vals
            best = max(pick_from)
            if qty_count > 0 and abs(best - qty_sum) < 0.01 and len(pick_from) > 1:
                alts = [v for v in pick_from if abs(v - qty_sum) >= 0.01]
                best = max(alts) if alts else best
            out["vendor_invoice_subtotal"] = best
        if out:
            return out
    return {}


def extract_vendor_invoice_fields(content_base64: str, file_name: str) -> dict[str, Any]:
    try:
        raw = base64.b64decode(content_base64, validate=True)
    except (ValueError, binascii.Error) as exc:
        raise ValueError("Invalid invoice file data. Re-upload the PDF, image, or Excel file.") from exc
    if len(raw) > 12 * 1024 * 1024:
        raise ValueError("File is too large (max 12 MB)")

    lower = file_name.lower()
    grid: dict[str, Any] = {}
    if lower.endswith((".xlsx", ".xlsm")):
        grid = _extract_from_excel_grid(raw)

    text = _text_from_bytes(raw, file_name)
    compact = re.sub(r"\s+", " ", text)
    inv_date = _find_date(compact)

    pair_qty, pair_amount = _find_qty_amount_pair(text)

    qty = _resolve_invoice_quantity(text, compact, grid.get("vendor_invoice_quantity"), pair_qty)

    exclude: set[float] = set()
    if qty is not None:
        exclude.add(float(qty))

    subtotal = grid.get("vendor_invoice_subtotal")
    if subtotal is None and pair_amount is not None:
        subtotal = pair_amount
    if subtotal is None:
        subtotal = _find_subtotal(compact, exclude)
    if subtotal is None:
        subtotal = _find_total_amount(compact, exclude)
    if subtotal is None:
        subtotal = _find_largest_money_fallback(compact, exclude)

    if qty is not None and subtotal is not None and abs(float(qty) - float(subtotal)) < 0.01:
        exclude.add(float(subtotal))
        subtotal = _find_total_amount(compact, exclude) or _find_largest_money_fallback(
            compact, exclude
        )

    inv_no = grid.get("vendor_invoice_number") or _find_invoice_number(compact)
    inv_date_str = grid.get("vendor_invoice_date")
    if not inv_date_str and inv_date:
        inv_date_str = inv_date.isoformat()

    return {
        "vendor_invoice_number": inv_no,
        "vendor_invoice_date": inv_date_str,
        "vendor_invoice_quantity": float(qty) if qty is not None else None,
        "vendor_invoice_subtotal": float(subtotal) if subtotal is not None else None,
    }
